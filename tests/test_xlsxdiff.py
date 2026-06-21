"""Tests for bizconnect.xlsxdiff + xlsxverify.

Core structural/alignment logic is driven on synthetic SheetViews built from
{(row, col): value} maps -- openpyxl cannot compute formula values, so a
workbook it writes has no cached values, and the alignment paths depend on them.
Synthetic views give deterministic control. Real .xlsx round-trips cover the
loader, number formats, and end-to-end emit/verify.
"""
from __future__ import annotations

import json

import openpyxl
import pytest
from openpyxl.worksheet.formula import ArrayFormula

from bizconnect import xlsxdiff as xd
from bizconnect import xlsxverify as xv
from bizconnect.xlsxdiff import SheetView, diff_sheet


def _sv(cells, formats=None):
    return SheetView(dict(cells), dict(cells), formats or {})


# --------------------------------------------------------------------------- #
# units / display rendering                                                   #
# --------------------------------------------------------------------------- #

def test_classify_unit():
    assert xd.classify_unit("0.0%")[0] == "percent"
    assert xd.classify_unit("#,##0")[0] == "integer"
    assert xd.classify_unit('"£"#,##0')[:2] == ("currency", "£")
    assert xd.classify_unit("General")[0] == "unknown"
    assert xd.classify_unit(None)[0] == "unknown"
    assert xd.classify_unit("yyyy-mm-dd")[0] == "date"


def test_render_display_formats():
    assert xd.render_display(0.175, "0.0%") == "17.5%"
    assert xd.render_display(0.085551, "0.00%") == "8.56%"
    assert xd.render_display(1234567.89, "#,##0") == "1,234,568"
    assert xd.render_display(1234.5, "#,##0") == "1,235"            # Excel half-up
    assert xd.render_display(1234.5, '"£"#,##0') == "£1,235"
    assert xd.render_display(-1234.5, '"£"#,##0;("£"#,##0)') == "(£1,235)"
    assert xd.render_display(None) == "(empty)"


def test_render_display_unknown_falls_back_not_wrong_unit():
    # an unrecognised format must NOT assert a unit
    assert xd.render_display(0.5, "General") in ("0.5", "0.5")
    assert xd.classify_unit("[DBNum1]")[0] == "unknown"


# --------------------------------------------------------------------------- #
# numeric hardening / helpers                                                  #
# --------------------------------------------------------------------------- #

def test_round_sig_handles_nonfinite_and_errors():
    for bad in (float("inf"), float("-inf"), float("nan"), 1e308, 1e-308, -0.0):
        xd._round_sig(bad)                              # no exception
    sv = _sv({(1, 1): "#DIV/0!", (1, 2): "#REF!"})
    sv.row_signature(1)                                 # error strings as literals, no crash


def test_values_differ_tolerance():
    assert not xd._values_differ(1.0, 1.0 + 1e-12)
    assert xd._values_differ(1.0, 1.1)
    assert not xd._values_differ(float("nan"), float("nan"))   # repr-equal
    assert xd._values_differ(None, 0)


def test_deshift_and_canon():
    assert xd._deshift("=A1", "A2", "A3") == "=A2"
    assert xd._deshift("=A1", "A2", "A2") == "=A1"
    assert xd.canon_formula("=sum( A1 , B1 )") == "=SUM(A1,B1)"


def test_is_formula_arrayformula():
    af = ArrayFormula(ref="A1", text="=SUM(B1:B2)")
    assert xd.is_formula(af) and xd.formula_text(af) == "=SUM(B1:B2)"


# --------------------------------------------------------------------------- #
# structural diff (synthetic)                                                  #
# --------------------------------------------------------------------------- #

def test_input_change():
    a = _sv({(1, 1): "Revenue", (1, 2): 100})
    b = _sv({(1, 1): "Revenue", (1, 2): 120})
    ins, dele, moved, changed, *_ = diff_sheet(a, b)
    assert ins == [] and dele == [] and moved == []
    assert changed[0][2] == [{"old_col": 2, "new_col": 2, "old": 100, "new": 120}]


def test_inserted_and_deleted_rows():
    a = _sv({(1, 1): "A", (2, 1): "B"})
    b = _sv({(1, 1): "A", (2, 1): "NEW", (3, 1): "B"})
    ins, dele, moved, *_ = diff_sheet(a, b)
    assert ins == [2] and dele == [] and moved == []
    ins, dele, moved, *_ = diff_sheet(b, a)
    assert dele == [2] and ins == [] and moved == []


def test_moved_row_and_label_fallback():
    a = _sv({(1, 1): "X", (2, 1): "Mover", (2, 2): 5, (3, 1): "Y", (4, 1): "Z"})
    b = _sv({(1, 1): "X", (2, 1): "Y", (3, 1): "Z", (4, 1): "Mover", (4, 2): 5})
    _i, _d, moved, *_ = diff_sheet(a, b)
    assert moved == [(2, 4)]
    # moved AND value rippled -> recovered by unique label
    b2 = _sv({(1, 1): "X", (2, 1): "Y", (3, 1): "Z", (4, 1): "Mover", (4, 2): 9})
    _i, _d, moved2, *_ = diff_sheet(a, b2)
    assert moved2 == [(2, 4)]


def test_block_move_collapse():
    # A 3-row block jumps PAST a longer stable region (A..H). difflib keeps the
    # longer region as 'equal' and sees the block as delete+insert; _match_moves
    # re-pairs the three -> one collapsed block move.
    stable = list("ABCDEFGH")
    old = {(1, 1): "Top"}
    for i, lab in enumerate(stable, start=2):
        old[(i, 1)] = lab; old[(i, 2)] = i
    for k, lab in enumerate(["BLK1", "BLK2", "BLK3"]):
        old[(10 + k, 1)] = lab; old[(10 + k, 2)] = 100 * (k + 1)
    new = {(1, 1): "Top"}
    for k, lab in enumerate(["BLK1", "BLK2", "BLK3"]):
        new[(2 + k, 1)] = lab; new[(2 + k, 2)] = 100 * (k + 1)
    for i, lab in enumerate(stable, start=5):
        new[(i, 1)] = lab; new[(i, 2)] = i - 3
    a, b = _sv(old), _sv(new)
    _i, _d, moved, changed, _ca, blocks, _cos = diff_sheet(a, b)
    assert len(blocks) == 1 and blocks[0]["n_rows"] == 3 and blocks[0]["delta"] == -8


def test_column_insert_is_one_fact_not_a_flood():
    # THE core fix: insert a column; assert one inserted column + row alignment
    # is NOT corrupted (no spurious value/input changes).
    old = {(1, 1): "Item", (1, 2): "2024", (1, 3): "2025",
           (2, 1): "Rev", (2, 2): 100, (2, 3): 110,
           (3, 1): "Cost", (3, 2): 40, (3, 3): 44}
    new = {(1, 1): "Item", (1, 2): "2024", (1, 3): "2025H1", (1, 4): "2025",
           (2, 1): "Rev", (2, 2): 100, (2, 3): 50, (2, 4): 110,
           (3, 1): "Cost", (3, 2): 40, (3, 3): 20, (3, 4): 44}
    a, b = _sv(old), _sv(new)
    ins, dele, moved, changed, ca, *_ = diff_sheet(a, b)
    assert [c for c in ca.inserted] == [3]              # the 2025H1 column
    assert ca.old_to_new == {1: 1, 2: 2, 3: 4}
    assert changed == []                                # Rev/Cost not falsely flagged


def test_pure_row_shift_does_not_flag_formula():
    a = SheetView({(1, 1): "Top", (2, 1): "=A1"}, {(1, 1): "Top", (2, 1): 10})
    b = SheetView({(1, 1): "Top", (2, 1): "Ins", (3, 1): "=A2"},
                  {(1, 1): "Top", (2, 1): "Ins", (3, 1): 10})
    ins, _d, _m, changed, *_ = diff_sheet(a, b, want_formulas=True)
    assert ins == [2]
    assert all(not c[4] for c in changed)              # no formula change reported


def test_value_preserving_formula_shift_is_suppressed():
    # formula text changed but cached value identical + deshift clean -> cosmetic
    a = SheetView({(1, 1): "X", (1, 2): "=Sheet2!A1"}, {(1, 1): "X", (1, 2): 7})
    b = SheetView({(1, 1): "X", (1, 2): "=Sheet2!A2"}, {(1, 1): "X", (1, 2): 7})
    _i, _d, _m, changed, _ca, _b, cosmetic = diff_sheet(a, b, want_formulas=True)
    assert cosmetic == 1
    assert all(not c[4] for c in changed)


# --------------------------------------------------------------------------- #
# roles / headlines                                                            #
# --------------------------------------------------------------------------- #

def test_role_classification():
    assert _sv({(1, 1): "Total Revenue", (1, 2): "=SUM(B2:B5)"}).classify_role(1)[0] == "output"
    assert _sv({(1, 1): "Total Costs", (1, 2): "=SUM(B2:B5)"}).classify_role(1)[0] == "total"
    assert _sv({(1, 1): "Headcount", (1, 2): 42}).classify_role(1)[0] == "input"


# --------------------------------------------------------------------------- #
# robust loader                                                                #
# --------------------------------------------------------------------------- #

@pytest.mark.parametrize("name,data,code", [
    ("corrupt.xlsx", b"not a zip at all", "CORRUPT"),
    ("legacy.xlsb", b"PK\x03\x04whatever", "LEGACY_FORMAT"),
    ("ole2.xlsx", b"\xd0\xcf\x11\xe0junk", "LEGACY_FORMAT"),
    ("empty.xlsx", b"", "CORRUPT"),
])
def test_loader_typed_errors(tmp_path, name, data, code):
    p = tmp_path / name
    p.write_bytes(data)
    with pytest.raises(xd.WorkbookLoadError) as e:
        xd.load_book(p)
    assert e.value.code == code


def test_compare_never_throws_on_bad_file(tmp_path):
    good = tmp_path / "good.xlsx"
    wb = openpyxl.Workbook(); wb.active["A1"] = "x"; wb.save(good)
    bad = tmp_path / "bad.xlsx"; bad.write_bytes(b"garbage")
    wd = xd.compare(good, bad)
    assert wd.status == "aborted"
    j = xd.emit_json(wd)
    assert j["status"] == "aborted" and j["diagnostics"]["errors"]


# --------------------------------------------------------------------------- #
# end-to-end on real workbooks                                                 #
# --------------------------------------------------------------------------- #

def _make(path, rows):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Summary"
    for (r, c), (val, nf) in rows.items():
        cell = ws.cell(row=r, column=c, value=val)
        if nf:
            cell.number_format = nf
    wb.save(path)


def test_emit_json_headlines_and_determinism(tmp_path):
    old = tmp_path / "old.xlsx"
    new = tmp_path / "new.xlsx"
    base = {
        (1, 1): ("Metric", None), (1, 2): ("FY24", None),
        (2, 1): ("Total Revenue", None), (2, 2): (100_000_000, "#,##0"),
        (3, 1): ("EBITDA Margin", None), (3, 2): (0.0806, "0.0%")}
    _make(old, base)
    upd = dict(base)
    upd[(2, 2)] = (105_500_000, "#,##0")
    upd[(3, 2)] = (0.0855, "0.0%")
    _make(new, upd)

    wd = xd.compare(old, new)
    j = xd.emit_json(wd)
    labels = {h["label"]: h for h in j["headline_metrics"]}
    assert "Total Revenue" in labels
    assert labels["Total Revenue"]["delta_display"] == "+5.5%"
    assert labels["Total Revenue"]["new_display"] == "105,500,000"
    assert labels["EBITDA Margin"]["delta_display"] == "+0.49pp"
    assert labels["EBITDA Margin"]["new_display"] == "8.6%"

    # byte-stable across runs
    s1 = json.dumps(xd.emit_json(xd.compare(old, new)), ensure_ascii=True)
    s2 = json.dumps(xd.emit_json(xd.compare(old, new)), ensure_ascii=True)
    assert s1 == s2
    assert j["diff_run_id"].startswith("wbdiff2-")


def test_named_ranges_and_rename(tmp_path):
    from openpyxl.workbook.defined_name import DefinedName
    old = tmp_path / "o.xlsx"; new = tmp_path / "n.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Model"
    for i, lab in enumerate(["Revenue", "Costs", "EBITDA", "Tax", "Net", "Cash"], 1):
        ws.cell(i, 1, lab); ws.cell(i, 2, i * 10)
    wb.defined_names.add(DefinedName("WACC", attr_text="Model!$B$1"))
    wb.save(old)
    wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.title = "Model v2"   # renamed
    for i, lab in enumerate(["Revenue", "Costs", "EBITDA", "Tax", "Net", "Cash"], 1):
        ws2.cell(i, 1, lab); ws2.cell(i, 2, i * 10)
    wb2.defined_names.add(DefinedName("Tax_Rate", attr_text="'Model v2'!$B$4"))
    wb2.save(new)
    j = xd.emit_json(xd.compare(old, new))
    assert any(r["old"] == "Model" and r["new"] == "Model v2" for r in j["renamed_sheets"])
    names = j["named_ranges"]
    assert any(n["name"] == "Tax_Rate" for n in names["added"])
    assert any(n["name"] == "WACC" for n in names["removed"])


def test_markdown_injection_escaped(tmp_path):
    old = tmp_path / "o.xlsx"; new = tmp_path / "n.xlsx"
    _make(old, {(1, 1): ("Item", None), (2, 1): ("safe", None), (2, 2): (1, None)})
    _make(new, {(1, 1): ("Item", None), (2, 1): ("| 999 | `x`", None), (2, 2): (2, None)})
    md = xd.render_markdown(xd.compare(old, new), values=True)
    assert "| 999 |" not in md or "\\|" in md            # pipe neutralised


# --------------------------------------------------------------------------- #
# verifier                                                                     #
# --------------------------------------------------------------------------- #

def _mini_diff():
    return {
        "diff_run_id": "wbdiff2-abc123",
        "headline_metrics": [
            {"id": "H01", "label": "Revenue", "old_display": "$100", "new_display": "$105.5m",
             "old_raw": 100, "new_raw": 105_500_000, "delta_display": "+5.5%", "delta_pct": 0.055}],
        "causal_links": [
            {"id": "C01", "cause_label": "Churn", "effect_headline_ids": ["H01"],
             "path_proven": True, "confidence": "high"},
            {"id": "C02", "cause_label": "FX", "effect_headline_ids": ["H01"],
             "path_proven": False, "confidence": "low"}],
        "sheets": [{"name": "S", "facts": [
            {"id": "F0001", "type": "input", "new_display": "5.0%", "old_display": "2.0%",
             "old_raw": 0.05, "new_raw": 0.02, "delta_display": "-3.00pp"}]}],
        "totals": {"rows_inserted": 5},
    }


def test_verifier_pass():
    n = "---\ndiff_run_id: wbdiff2-abc123\n---\nRevenue rose +5.5% {H01}. Churn fell to 2.0% {F0001}."
    assert xv.verify(n, _mini_diff())["status"] == "PASS"


def test_verifier_flags_fabrication_and_miscite():
    d = _mini_diff()
    assert xv.verify("---\ndiff_run_id: wbdiff2-abc123\n---\nRevenue $999,999,999 {H01}.", d)["status"] == "FAIL"
    assert xv.verify("---\ndiff_run_id: wbdiff2-abc123\n---\nUp +5.5% {F9999}.", d)["status"] == "FAIL"


def test_verifier_runid_and_causal_and_failclosed():
    d = _mini_diff()
    assert xv.verify("---\ndiff_run_id: wbdiff2-WRONG\n---\nx +5.5% {H01}", d)["status"] == "FAIL"
    bad_cause = "---\ndiff_run_id: wbdiff2-abc123\n---\nFX drove revenue +5.5% {C02} {H01}."
    assert xv.verify(bad_cause, d)["status"] == "FAIL"          # weak causal edge
    assert xv.verify("no front matter here", d)["status"] == "FAIL"   # fail-closed


# ---- regression tests for the review fixes ----

def test_classify_unit_date_locale_and_quoted_currency():
    assert xd.classify_unit("[$-409]m/d/yyyy")[0] == "date"          # locale $, not currency
    assert xd.classify_unit('[$-409]#,##0')[0] in ("integer", "number")  # locale only
    assert xd.classify_unit('"£"#,##0')[:2] == ("currency", "£")     # quoted-literal currency
    assert xd.classify_unit('[$£-809] #,##0')[:2] == ("currency", "£")


def test_delta_negative_base_has_no_misleading_percent():
    d = xd._delta(-244_891_153, -274_959_995, "currency", "currency")
    assert "delta_pct" not in d and "delta_display" not in d         # abs only; caller renders
    assert d["delta_abs"] < 0


def test_as_of_col_prefers_total_over_month_with_year():
    view = SheetView(
        {(1, 1): "Item", (1, 2): "FY2025 Total", (1, 3): "Jan 2026",
         (2, 1): "Revenue", (2, 2): 1000, (2, 3): 80},
        {(1, 1): "Item", (1, 2): "FY2025 Total", (1, 3): "Jan 2026",
         (2, 1): "Revenue", (2, 2): 1000, (2, 3): 80})
    assert xd._as_of_col(view, 2, [2, 3]) == 2                       # the FY total, not Jan-2026


def test_role_excludes_cost_lines_from_output():
    assert _sv({(1, 1): "Cost of Sales", (1, 2): 5}).classify_role(1)[0] != "output"


def test_verifier_bare_number_and_unit_dimension():
    d = _mini_diff()
    # a fabricated BARE number must FAIL (no comma/symbol must not let it escape)
    assert xv.verify("---\ndiff_run_id: wbdiff2-abc123\n---\nRevenue was 777777777 this year.", d)["status"] == "FAIL"
    # a percent token may not be grounded by an absolute fact value of the same digits
    assert xv.verify("---\ndiff_run_id: wbdiff2-abc123\n---\nMargin hit 105500000%.", d)["status"] == "FAIL"
    # front-matter bypass closed: a prose line starting with 'schema_version' is still checked
    assert xv.verify("---\ndiff_run_id: wbdiff2-abc123\n---\nschema_version fabrications cost $888,888,888.", d)["status"] == "FAIL"
