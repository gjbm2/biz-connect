"""Structural, human-readable diff of two .xlsx workbooks (engine).

Financial models are full of formulas that reference other cells. Insert or
delete a single row/column and Excel rewrites thousands of references -- a naive
cell-by-cell diff then drowns in shift noise. This engine diffs *structurally*:

  1. Align COLUMNS (1D), then ROWS (1D, on a column-shift-invariant signature),
     so an inserted row/column is one structural fact, not thousands of shifts.
  2. Classify each row's role (input / total / output / ...) and capture number
     formats, so values read the way the model shows them (0.175 -> "17.5%").
  3. Emit a deterministic JSON *fact graph* -- the single source of truth. Every
     change is an atomic fact with a stable id (F0001) carrying old_raw (exact)
     AND old_display (formatted), unit, role, materiality, and cause/effect tier.
     Headline metrics (Revenue/EBITDA/margin deltas) are extracted into a closed,
     citable set. A separate verifier (xlsxverify) checks a narrative against it.

Design notes (see docs/xlsxdiff-v2-spec.json):
  * "A plain-but-correct diff beats a rich-but-wrong one." Risky inferences
    (column alignment, causal links) are confidence-tagged with safe fallbacks.
  * compare() NEVER raises to the caller: load/format/size failures come back as
    status='aborted'|'degraded' with diagnostics, so the skill can relay them.
  * JSON is byte-stable across runs (deterministic ids + content-hash run id),
    so the verifier can bind a narrative to exactly one diff run.

Public API:
    compare(old, new, *, formulas=False, want_graph=True) -> WorkbookDiff
    emit_json(wd) -> dict                 # the ground-truth fact graph
    render_markdown(wd, *, values, formulas, max_rows) -> str   # capped preview
    diff_to_markdown(...) / diff_to_json(...)                    # convenience
"""
from __future__ import annotations

import datetime as _dt
import hashlib
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from difflib import SequenceMatcher
from math import floor, isfinite, log10
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.formula import ArrayFormula

SCHEMA_VERSION = "wbdiff/2"
ENGINE_VERSION = "0.7.0"

# Hard caps (per the spec: bound COMPUTATION, not just rendering).
MAX_ROWS = 200_000
MAX_COLS = 1_024
MAX_RANGE_EXPAND = 20_000      # cap on expanding a referenced range in the graph
EPS_REL = 1e-9                 # unified relative tolerance (signature + value diff)
EPS_FLOOR = 1e-9

EXCEL_ERRORS = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!", "#SPILL!", "#CALC!"}


# --------------------------------------------------------------------------- #
# Typed load failure                                                           #
# --------------------------------------------------------------------------- #

class WorkbookLoadError(Exception):
    def __init__(self, code, path, detail):
        self.code = code
        self.path = str(path)
        self.detail = detail
        super().__init__(f"{code}: {detail}")


def _sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _validate_format(path):
    """Suffix allowlist + magic-byte sniff. Raises WorkbookLoadError on a clear
    non-xlsx (legacy/encrypted/corrupt) so we never hand openpyxl a doomed read."""
    p = Path(path)
    if not p.exists():
        raise WorkbookLoadError("UNREADABLE", p, f"file not found: {p}")
    suffix = p.suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        if suffix in (".xls", ".xlsb"):
            raise WorkbookLoadError("LEGACY_FORMAT", p,
                f"openpyxl cannot read {suffix}; save as .xlsx and re-run.")
        # not obviously excel -- still sniff before rejecting
    try:
        with open(p, "rb") as fh:
            magic = fh.read(8)
    except OSError as e:
        raise WorkbookLoadError("UNREADABLE", p, f"cannot read file: {e}")
    if magic[:4] == b"PK\x03\x04":
        return                                   # zip container -> probably xlsx
    if magic[:4] == b"\xd0\xcf\x11\xe0":
        raise WorkbookLoadError("LEGACY_FORMAT", p,
            "file is an OLE2 document (legacy .xls or an encrypted .xlsx); "
            "open it in Excel and Save As .xlsx (unencrypted), then re-run.")
    raise WorkbookLoadError("CORRUPT", p,
        f"not a readable .xlsx (unexpected header {magic[:4]!r}).")


def _zip_bomb_guard(path):
    """Reject absurd uncompressed expansion before openpyxl parses it."""
    try:
        with zipfile.ZipFile(path) as zf:
            total = sum(i.file_size for i in zf.infolist())
            if total > 2_000_000_000:            # 2 GB uncompressed
                raise WorkbookLoadError("CORRUPT", path,
                    f"uncompressed size {total} bytes exceeds safety cap.")
    except zipfile.BadZipFile:
        raise WorkbookLoadError("CORRUPT", path, "not a valid zip/xlsx container.")


# --------------------------------------------------------------------------- #
# Number formats / units / display                                            #
# --------------------------------------------------------------------------- #

_CURRENCY = ("$", "£", "€", "¥", "₹", "₩")
_SCALE_SUFFIX = {0: "", 1: "k", 2: "m", 3: "bn", 4: "tn"}


def _decimals(nf):
    m = re.search(r"\.([0#]+)", nf)
    return len(m.group(1)) if m else 0


def _scale_commas(nf):
    m = re.search(r"[0#](,+)(?=[^0#,]*$)", nf)
    return len(m.group(1)) if m else 0


def classify_unit(number_format):
    """Return (unit_kind, currency_symbol, scale) for an Excel number format.

    unit_kind in {percent, currency, scaled, integer, number, date, text, unknown}.
    Conservative: an unrecognised format yields 'unknown' (never a wrong unit)."""
    if not number_format or number_format in ("General",):
        return ("unknown", None, 0)
    s = number_format.split(";")[0]
    if "%" in s:
        return ("percent", None, 0)
    sym = next((c for c in _CURRENCY if c in s), None)
    scale = _scale_commas(s)
    if sym:
        return ("currency", sym, scale)
    if _is_date_format(s):
        return ("date", None, 0)
    if scale:
        return ("scaled", None, scale)
    if re.search(r"[0#]", s):
        return ("integer" if "." not in s else "number", None, 0)
    return ("unknown", None, 0)


def _is_date_format(s):
    core = re.sub(r'\[[^\]]*\]|"[^"]*"|\\.', "", s)
    return bool(re.search(r"[ymdhs]", core, re.I)) and not re.search(r"[0#]", core)


def _round_hu(x, dec):
    try:
        q = Decimal(1).scaleb(-dec)
        return Decimal(repr(x)).quantize(q, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal(0)


def render_display(raw, number_format=None):
    """Render a raw value the way the model shows it (Excel half-up rounding)."""
    if raw is None:
        return "(empty)"
    if isinstance(raw, bool):
        return str(raw)
    if isinstance(raw, (_dt.datetime, _dt.date)):
        if isinstance(raw, _dt.datetime) and raw.time() == _dt.time(0, 0):
            return raw.date().isoformat()
        return raw.isoformat()
    if isinstance(raw, str):
        return raw.replace("\n", " ").strip()
    if not isinstance(raw, (int, float)):
        return str(raw)
    if isinstance(raw, float) and not isfinite(raw):
        return str(raw)
    kind, sym, scale = classify_unit(number_format)
    pos = (number_format or "General").split(";")[0]
    if kind == "percent":
        dec = _decimals(pos)
        d = _round_hu(raw * 100, dec)
        return f"{d:.{dec}f}%"
    if kind in ("currency", "scaled"):
        dec = _decimals(pos)
        val = raw / (1000 ** scale) if scale else raw
        d = _round_hu(abs(val), dec)
        thousands = bool(re.search(r"#,#|0,0|#,##0", pos))
        body = f"{d:,.{dec}f}" if thousands else f"{d:.{dec}f}"
        out = f"{sym or ''}{body}{_SCALE_SUFFIX.get(scale, '')}"
        return f"({out})" if (raw < 0 and "(" in (number_format or "")) else (f"-{out}" if raw < 0 else out)
    if kind in ("integer", "number"):
        dec = _decimals(pos)
        thousands = bool(re.search(r"#,#|0,0|#,##0", pos))
        d = _round_hu(raw, dec)
        return f"{d:,.{dec}f}" if thousands else f"{d:.{dec}f}"
    # unknown / general -> compact generic
    return _fmt_generic(raw)


def _fmt_generic(v):
    if isinstance(v, float):
        if isfinite(v) and v == int(v):
            return f"{int(v):,}"
        return f"{v:,.10g}"
    if isinstance(v, int):
        return f"{v:,}"
    return str(v)


# Back-compat thin alias used by older callers/tests.
def fmt(value):
    return render_display(value, None)


# --------------------------------------------------------------------------- #
# Cell helpers                                                                 #
# --------------------------------------------------------------------------- #

def addr(r, c):
    return f"{get_column_letter(c)}{r}"


def is_formula(v):
    return isinstance(v, ArrayFormula) or (isinstance(v, str) and v.startswith("="))


def formula_text(v):
    if isinstance(v, ArrayFormula):
        return v.text
    return v


def _is_error(v):
    return isinstance(v, str) and v in EXCEL_ERRORS


def canon_formula(text):
    """Conservative, provably-equivalent canonicalisation for benign-resave noise:
    trim, collapse inter-token whitespace, uppercase function names. Anything
    ambiguous is left intact (we'd rather over-report than mask a real change)."""
    if not isinstance(text, str):
        return text
    t = text.strip()
    # collapse whitespace that is not inside a quoted string
    out, instr, i = [], False, 0
    while i < len(t):
        ch = t[i]
        if ch == '"':
            instr = not instr
            out.append(ch)
        elif ch.isspace() and not instr:
            if out and not out[-1].isspace():
                out.append(" ")
        else:
            out.append(ch)
        i += 1
    s = "".join(out)
    # uppercase function names (identifier immediately followed by '(')
    s = re.sub(r"([A-Za-z_][A-Za-z0-9_.]*)\s*\(",
               lambda m: m.group(1).upper() + "(", s)
    # strip spaces adjacent to delimiters, but never inside string literals
    parts = re.split(r'("[^"]*")', s)
    for i in range(0, len(parts), 2):
        parts[i] = re.sub(r"\s*([(),:])\s*", r"\1", parts[i])
    return "".join(parts)


def _values_differ(a, b):
    if a is None and b is None:
        return False
    if isinstance(a, bool) or isinstance(b, bool):
        return a != b
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if isinstance(a, float) and not isfinite(a):
            return repr(a) != repr(b)
        if isinstance(b, float) and not isfinite(b):
            return repr(a) != repr(b)
        return abs(a - b) > max(EPS_FLOOR, EPS_REL * max(1.0, abs(a), abs(b)))
    return a != b


def _deshift(formula, old_addr, new_addr):
    if old_addr == new_addr:
        return formula
    try:
        return Translator(formula, origin=old_addr).translate_formula(new_addr)
    except Exception:
        return None             # signal failure (deshift_ok = False)


def _round_sig(v, sig=6):
    if isinstance(v, float):
        if not isfinite(v):
            return ("nonfinite", repr(v))
        if v == 0:
            return 0.0
        d = sig - 1 - floor(log10(abs(v)))
        return round(v, d)
    return v


# --------------------------------------------------------------------------- #
# Markdown safety                                                              #
# --------------------------------------------------------------------------- #

_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f‪-‮⁦-⁩]")


def md_escape(s):
    """Neutralise workbook text so a cell can't inject table/markdown structure
    into the report the agent grounds on."""
    if not isinstance(s, str):
        s = render_display(s, None)
    s = _CTRL.sub("", s)
    s = s.replace("\\", "\\\\").replace("|", "\\|").replace("`", "\\`")
    s = s.replace("\r", " ").replace("\n", " ")
    # cell text is always interpolated mid-line (table cells / after "- **"), so
    # block markers (#, -, >) can't trigger; only neutralise the table/code chars.
    return s.strip()


# --------------------------------------------------------------------------- #
# Loading                                                                      #
# --------------------------------------------------------------------------- #

@dataclass
class LoadedBook:
    name: str
    sha256: str
    order: list
    content: dict        # {sheet: {(r,c): formula/literal}}
    values: dict         # {sheet: {(r,c): cached value}}
    formats: dict        # {sheet: {(r,c): number_format}}
    defined_names: dict  # {name: refers_to}
    coverage: dict       # {sheet: cached-value coverage 0..1}
    fmt_cells: int = 0
    fmt_known: int = 0


def _iter_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                yield cell


def load_book(path):
    """Load one workbook -> LoadedBook. Raises WorkbookLoadError on failure."""
    _validate_format(path)
    _zip_bomb_guard(path)
    try:
        wb_f = openpyxl.load_workbook(path, data_only=False, read_only=True)
        wb_v = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, OSError, ValueError) as e:
        msg = str(e)
        code = "CORRUPT"
        if "password" in msg.lower() or "encrypt" in msg.lower():
            code = "ENCRYPTED"
        raise WorkbookLoadError(code, path, f"openpyxl could not open the workbook: {msg}")

    order = list(wb_f.sheetnames)
    content, formats, coverage = {}, {}, {}
    fmt_cells = fmt_known = 0
    for name in order:
        ws = wb_f[name]
        cmap, fmap = {}, {}
        formula_cells = formula_with_cache = 0
        for cell in _iter_cells(ws):
            key = (cell.row, cell.column)
            cmap[key] = cell.value
            nf = getattr(cell, "number_format", None)
            if nf and nf != "General":
                fmap[key] = nf
                fmt_known += 1
            fmt_cells += 1
        content[name] = cmap
        formats[name] = fmap

    values = {}
    for name in wb_v.sheetnames:
        ws = wb_v[name]
        vmap = {}
        for cell in _iter_cells(ws):
            vmap[(cell.row, cell.column)] = cell.value
        values[name] = vmap

    # cached-value coverage: of cells that are formulas in content, how many have
    # a non-None cached value?
    for name in order:
        cmap = content.get(name, {})
        vmap = values.get(name, {})
        fcells = [k for k, v in cmap.items() if is_formula(v)]
        if not fcells:
            coverage[name] = 1.0
        else:
            have = sum(1 for k in fcells if vmap.get(k) is not None)
            coverage[name] = have / len(fcells)

    dn = {}
    try:
        for k in wb_f.defined_names:
            d = wb_f.defined_names[k]
            dn[k] = getattr(d, "attr_text", None) or getattr(d, "value", None)
    except Exception:
        dn = {}

    wb_f.close()
    wb_v.close()
    return LoadedBook(Path(path).name, _sha256(path), order, content, values,
                      formats, dn, coverage, fmt_cells, fmt_known)


# --------------------------------------------------------------------------- #
# Sheet view                                                                   #
# --------------------------------------------------------------------------- #

_OUTPUT_LEX = re.compile(
    r"\b(net income|net profit|ebitda|ebit|free cash ?flow|fcf|npv|irr|revenue|"
    r"sales|turnover|gross profit|operating profit|profit before tax|pbt|"
    r"net cash|contribution|gross margin|ebitda margin|cac|ltv|arpu)\b", re.I)
_TOTAL_LEX = re.compile(r"\b(total|subtotal|sum of|grand total)\b", re.I)
_MARGIN_LEX = re.compile(r"\b(margin|%|growth|rate)\b", re.I)


class SheetView:
    """Row/column-indexed view built from content + value (+ format) maps.

    Constructible directly from dicts for testing:
        SheetView({(1,1): "=A2"}, {(1,1): 42})
    """

    def __init__(self, content, values, formats=None):
        self.content = dict(content)
        self.values = dict(values)
        self.formats = dict(formats or {})
        keys = set(self.content) | set(self.values)
        self._rows_set = sorted({r for (r, _c) in keys})
        self._cols_set = sorted({c for (_r, c) in keys})
        self.min_row = self._rows_set[0] if self._rows_set else 1
        self.max_row = self._rows_set[-1] if self._rows_set else 0
        self.max_col = self._cols_set[-1] if self._cols_set else 0
        self._rc_content = self._index_by_row(self.content)
        self._rc_values = self._index_by_row(self.values)
        self._cc_content = self._index_by_col(self.content)
        self._cc_values = self._index_by_col(self.values)

    @staticmethod
    def _index_by_row(m):
        out = {}
        for (r, c), v in m.items():
            out.setdefault(r, {})[c] = v
        return out

    @staticmethod
    def _index_by_col(m):
        out = {}
        for (r, c), v in m.items():
            out.setdefault(c, {})[r] = v
        return out

    def populated_rows(self):
        return self._rows_set

    def populated_cols(self):
        return self._cols_set

    def header_row(self):
        """First row with >=2 text cells (a likely header), else min_row."""
        for r in self._rows_set[:25]:
            texts = sum(1 for v in self._rc_values.get(r, {}).values()
                        if isinstance(v, str) and v.strip())
            if texts >= 2:
                return r
        return self.min_row

    def col_header(self, c):
        hv = self._rc_values.get(self.header_row(), {}).get(c)
        if isinstance(hv, str) and hv.strip():
            return hv.strip()
        cv = self._rc_content.get(self.header_row(), {}).get(c)
        if isinstance(cv, str) and cv.strip() and not is_formula(cv):
            return cv.strip()
        return ""

    def label(self, r, max_col=6):
        rc = self._rc_values.get(r, {})
        cc = self._rc_content.get(r, {})
        for c in range(1, max_col + 1):
            v = rc.get(c)
            if isinstance(v, str) and v.strip():
                return v.strip()
        for c in range(1, max_col + 1):
            v = cc.get(c)
            if isinstance(v, str) and v.strip() and not is_formula(v):
                return v.strip()
        return None

    def row_signature(self, r, cols=None):
        """Shift-invariant row identity. If `cols` (common columns) is given,
        values are taken only at those columns so a column insert/delete does
        NOT perturb row identity."""
        label = (self.label(r) or "").lower()
        rc = self._rc_values.get(r, {})
        items = rc.items() if cols is None else ((c, rc[c]) for c in cols if c in rc)
        vsig = tuple(sorted((c, _round_sig(v)) for c, v in items
                            if not isinstance(v, str) or v.strip()))
        return (label, vsig)

    def col_signature(self, c):
        header = self.col_header(c).lower()
        cv = self._cc_values.get(c, {})
        vsig = tuple(sorted((r, _round_sig(v)) for r, v in cv.items()
                            if isinstance(v, (int, float)) and not isinstance(v, bool)
                            and isfinite(v) if True))
        return (header, vsig)

    def row_content(self, r):
        return self._rc_content.get(r, {})

    def row_values(self, r):
        return self._rc_values.get(r, {})

    def fmt_at(self, r, c):
        return self.formats.get((r, c))

    def is_blank(self, r):
        return not self._rc_content.get(r) and not self._rc_values.get(r)

    def classify_role(self, r):
        """(role, confidence) for a row. Conservative: corroboration required for
        'output'/'total' before they may drive a headline."""
        label = self.label(r) or ""
        content = self._rc_content.get(r, {})
        has_formula = any(is_formula(v) for v in content.values())
        has_literal_number = any(isinstance(v, (int, float)) and not isinstance(v, bool)
                                 for v in content.values())
        sum_shape = any(isinstance(v, str) and re.match(r"=\s*(SUM|SUBTOTAL)\b", v, re.I)
                        for v in content.values())
        if _OUTPUT_LEX.search(label):
            conf = "high" if (has_formula or sum_shape) else "medium"
            return ("output", conf)
        if _TOTAL_LEX.search(label) or sum_shape:
            return ("total", "high" if sum_shape else "medium")
        if has_formula:
            return ("intermediate", "medium")
        if has_literal_number:
            return ("input", "medium")
        if label:
            return ("header", "low")
        return ("spacer", "low")


# --------------------------------------------------------------------------- #
# Column alignment (1D, independent pass)                                      #
# --------------------------------------------------------------------------- #

@dataclass
class ColAlign:
    old_to_new: dict          # matched old col -> new col
    inserted: list            # new cols with no old match
    deleted: list             # old cols with no new match
    moved: list               # (old_col, new_col) relocations
    confidence: str           # high|medium|low
    method: str               # 'col-1d' | 'identity'


def align_columns(va: SheetView, vb: SheetView):
    cols_a, cols_b = va.populated_cols(), vb.populated_cols()
    # If headers are largely absent, column identity is unreliable -> identity map.
    headers_a = sum(1 for c in cols_a if va.col_header(c))
    headers_b = sum(1 for c in cols_b if vb.col_header(c))
    text_ratio = (headers_a + headers_b) / max(1, len(cols_a) + len(cols_b))

    if text_ratio < 0.34 or not cols_a or not cols_b:
        common = [c for c in cols_b if c in set(cols_a)]
        return ColAlign({c: c for c in common}, [], [], [], "low", "identity")

    sig_a = [va.col_signature(c) for c in cols_a]
    sig_b = [vb.col_signature(c) for c in cols_b]
    sm = SequenceMatcher(a=sig_a, b=sig_b, autojunk=False)
    old_to_new, inserted, deleted = {}, [], []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            for k in range(i2 - i1):
                old_to_new[cols_a[i1 + k]] = cols_b[j1 + k]
        elif tag == "replace":
            n = min(i2 - i1, j2 - j1)
            for k in range(n):
                old_to_new[cols_a[i1 + k]] = cols_b[j1 + k]
            inserted.extend(cols_b[j1 + n:j2])
            deleted.extend(cols_a[i1 + n:i2])
        elif tag == "insert":
            inserted.extend(cols_b[j1:j2])
        elif tag == "delete":
            deleted.extend(cols_a[i1:i2])

    # recover column moves by matching insert+delete with equal header
    moved = []
    del_by_hdr = defaultdict(list)
    for c in deleted:
        del_by_hdr[va.col_header(c).lower()].append(c)
    rem_ins, used = [], set()
    for nc in inserted:
        h = vb.col_header(nc).lower()
        cand = del_by_hdr.get(h)
        if h and cand and len(cand) == 1 and cand[0] not in used:
            oc = cand[0]
            used.add(oc)
            moved.append((oc, nc))
            old_to_new[oc] = nc
        else:
            rem_ins.append(nc)
    deleted = [c for c in deleted if c not in used]

    matched = len(old_to_new)
    conf = "high" if matched >= 0.8 * min(len(cols_a), len(cols_b)) else \
           "medium" if matched >= 0.5 * min(len(cols_a), len(cols_b)) else "low"
    if conf == "low":                                  # fall back to identity
        common = [c for c in cols_b if c in set(cols_a)]
        return ColAlign({c: c for c in common}, [], [], [], "low", "identity")
    return ColAlign(old_to_new, sorted(rem_ins), sorted(deleted), moved, conf, "col-1d")


# --------------------------------------------------------------------------- #
# Row diff over the aligned lattice                                           #
# --------------------------------------------------------------------------- #

def _diff_aligned_row(va, ra, vb, rb, col_map, want_formulas):
    """Compare one aligned row pair over matched columns.

    Returns (inputs, values, formulas, cosmetic_shifts). A formula whose text
    changed but whose CACHED VALUE did not (and which de-shifted cleanly) is
    almost certainly a benign reference renumber from a structural edit
    elsewhere -- it is suppressed (counted in cosmetic_shifts), not reported as
    a logic change. This kills the cross-sheet / insert-boundary shift flood."""
    inputs, values, formulas = [], [], []
    cosmetic = 0
    ca, cb = va.row_content(ra), vb.row_content(rb)
    xa, xb = va.row_values(ra), vb.row_values(rb)
    for oc, nc in col_map.items():
        a, b = ca.get(oc), cb.get(nc)
        af, bf = is_formula(a), is_formula(b)
        va_, vb_ = xa.get(oc), xb.get(nc)
        value_changed = _values_differ(va_, vb_)

        if not af and not bf:                          # literal/input cell
            if a != b and not (a is None and b is None):
                inputs.append({"old_col": oc, "new_col": nc, "old": a, "new": b})
        elif want_formulas:
            at = formula_text(a) if af else a
            bt = formula_text(b) if bf else b
            deshift = _deshift(at, addr(ra, oc), addr(rb, nc)) if af else at
            ok = deshift is not None
            if canon_formula(deshift if ok else at) != canon_formula(bt):
                if value_changed or not ok:
                    formulas.append({"old_col": oc, "new_col": nc,
                                     "old": at, "new": bt, "deshift_ok": ok})
                else:
                    cosmetic += 1                      # value-preserving renumber

        if value_changed:
            values.append({"old_col": oc, "new_col": nc, "old": va_, "new": vb_})
    return inputs, values, formulas, cosmetic


def diff_sheet(va: SheetView, vb: SheetView, want_formulas=False, col_align=None):
    """Diff one sheet. Returns (inserted, deleted, moved, changed, col_align).

    changed: list of (ra, rb, inputs, values, formulas).  Backwards compatible
    with the previous 4-tuple plus the column alignment as a 5th element."""
    if col_align is None:
        col_align = align_columns(va, vb)
    common_cols_old = list(col_align.old_to_new.keys())

    rows_a, rows_b = va.populated_rows(), vb.populated_rows()
    common_new_cols = set(col_align.old_to_new.values())
    sig_a = [va.row_signature(r, cols=common_cols_old) for r in rows_a]
    sig_b = [vb.row_signature(r, cols=common_new_cols) for r in rows_b]
    sm = SequenceMatcher(a=sig_a, b=sig_b, autojunk=False)

    inserted, deleted, changed = [], [], []
    cosmetic = [0]
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            for k in range(i2 - i1):
                _emit_row(va, rows_a[i1 + k], vb, rows_b[j1 + k],
                          col_align.old_to_new, want_formulas, changed, cosmetic)
        elif tag == "replace":
            n = min(i2 - i1, j2 - j1)
            for k in range(n):
                _emit_row(va, rows_a[i1 + k], vb, rows_b[j1 + k],
                          col_align.old_to_new, want_formulas, changed, cosmetic)
            for k in range(n, j2 - j1):
                rb = rows_b[j1 + k]
                if not vb.is_blank(rb):
                    inserted.append(rb)
            for k in range(n, i2 - i1):
                ra = rows_a[i1 + k]
                if not va.is_blank(ra):
                    deleted.append(ra)
        elif tag == "insert":
            inserted.extend(r for r in rows_b[j1:j2] if not vb.is_blank(r))
        elif tag == "delete":
            deleted.extend(r for r in rows_a[i1:i2] if not va.is_blank(r))

    moved, inserted, deleted = _match_moves(va, vb, inserted, deleted, common_cols_old,
                                            list(common_new_cols))
    moved_blocks = _collapse_blocks(moved)
    return inserted, deleted, moved, changed, col_align, moved_blocks, cosmetic[0]


def _emit_row(va, ra, vb, rb, col_map, want_formulas, changed, cosmetic):
    ins, val, fml, cos = _diff_aligned_row(va, ra, vb, rb, col_map, want_formulas)
    cosmetic[0] += cos
    if ins or val or fml:
        changed.append((ra, rb, ins, val, fml))


def _match_moves(va, vb, inserted, deleted, cols_old, cols_new):
    del_by_sig = defaultdict(list)
    for ra in deleted:
        del_by_sig[va.row_signature(ra, cols=cols_old)].append(ra)
    for v in del_by_sig.values():
        v.sort()
    moved, rem_inserted, used = [], [], set()
    for rb in sorted(inserted):
        cand = del_by_sig.get(vb.row_signature(rb, cols=cols_new))
        ra = None
        while cand:
            x = cand.pop(0)
            if x not in used:
                ra = x
                break
        if ra is not None:
            used.add(ra)
            moved.append((ra, rb))
        else:
            rem_inserted.append(rb)
    rem_deleted = [ra for ra in deleted if ra not in used]

    ins_labels = Counter(vb.label(rb) for rb in rem_inserted)
    del_labels = Counter(va.label(ra) for ra in rem_deleted)
    del_by_label = defaultdict(list)
    for ra in rem_deleted:
        del_by_label[va.label(ra)].append(ra)
    for v in del_by_label.values():
        v.sort()
    final_inserted, used2 = [], set()
    for rb in sorted(rem_inserted):
        lab = vb.label(rb)
        if lab and ins_labels[lab] == 1 and del_labels.get(lab) and len(del_by_label[lab]) == 1:
            ra = del_by_label[lab][0]
            if ra not in used2:
                moved.append((ra, rb))
                used2.add(ra)
                continue
        final_inserted.append(rb)
    final_deleted = [ra for ra in rem_deleted if ra not in used2]
    moved.sort(key=lambda t: (t[1], t[0]))
    return moved, sorted(final_inserted), sorted(final_deleted)


def _collapse_blocks(moved):
    """Collapse contiguous runs with constant delta into block moves."""
    if not moved:
        return []
    by_old = sorted(moved)
    blocks, run = [], [by_old[0]]
    for prev, cur in zip(by_old, by_old[1:]):
        same_delta = (cur[0] - prev[0] == 1) and (cur[1] - prev[1] == 1)
        if same_delta:
            run.append(cur)
        else:
            blocks.append(run)
            run = [cur]
    blocks.append(run)
    out = []
    for run in blocks:
        if len(run) >= 3:                              # only collapse real blocks
            out.append({"old_rows": [run[0][0], run[-1][0]],
                        "new_rows": [run[0][1], run[-1][1]],
                        "delta": run[0][1] - run[0][0], "n_rows": len(run),
                        "members": run})
    return out


# --------------------------------------------------------------------------- #
# Dependency graph + causal attribution (advisory)                            #
# --------------------------------------------------------------------------- #

def _cells_from_ref(val, cur_sheet, names):
    """Yield (sheet,row,col) cells named by one reference token, resolving a
    defined name and expanding (capped) ranges. Whole-row/column refs skipped."""
    val = val.strip()
    if names and val in names and names[val]:
        rt = names[val]
        rt = rt[1:] if rt.startswith("=") else rt
        for piece in rt.split(","):
            yield from _cells_from_ref(piece.strip(), cur_sheet, None)
        return
    sheet = cur_sheet
    if "!" in val:
        sp, _, val = val.rpartition("!")
        sheet = sp.strip().strip("'").replace("''", "'")
    val = val.replace("$", "")
    try:
        if ":" in val:
            a, _, b = val.partition(":")
            ca, ra = coordinate_from_string(a)
            cb, rb = coordinate_from_string(b)
            c1, c2 = sorted((column_index_from_string(ca), column_index_from_string(cb)))
            r1, r2 = sorted((ra, rb))
            if (r2 - r1 + 1) * (c2 - c1 + 1) > MAX_RANGE_EXPAND:
                return
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    yield (sheet, rr, cc)
        else:
            cc, rr = coordinate_from_string(val)
            yield (sheet, rr, column_index_from_string(cc))
    except Exception:
        return                                          # whole-col/row or unparseable -> skip


def _formula_refs(text, cur_sheet, names=None):
    """Yield precedent cells of a formula using openpyxl's tokenizer (robustly
    identifies RANGE operands, unlike a regex over the whole string)."""
    if not isinstance(text, str):
        return
    try:
        from openpyxl.formula.tokenizer import Tokenizer
        toks = Tokenizer(text).items
    except Exception:
        return
    for t in toks:
        if t.type == "OPERAND" and t.subtype == "RANGE":
            yield from _cells_from_ref(t.value, cur_sheet, names)


def build_dependents(books_content, names=None):
    """dependents[(sheet,r,c)] = set of cells whose formula references it.
    Built on the NEW workbook content (single coordinate space)."""
    deps = defaultdict(set)
    for sheet, cmap in books_content.items():
        for (r, c), v in cmap.items():
            if not is_formula(v):
                continue
            for ref in _formula_refs(formula_text(v), sheet, names):
                deps[ref].add((sheet, r, c))
    return deps


def reachable(deps, start, cap=200_000):
    seen, stack = set(), [start]
    while stack and len(seen) < cap:
        node = stack.pop()
        for d in deps.get(node, ()):  # cells depending on `node`
            if d not in seen:
                seen.add(d)
                stack.append(d)
    return seen


# --------------------------------------------------------------------------- #
# Fact / diff data model                                                       #
# --------------------------------------------------------------------------- #

@dataclass
class SheetDiff:
    name: str
    inserted: list = field(default_factory=list)
    deleted: list = field(default_factory=list)
    moved: list = field(default_factory=list)
    moved_blocks: list = field(default_factory=list)
    changed: list = field(default_factory=list)
    col_align: ColAlign = None
    cosmetic_formula_shifts: int = 0

    @property
    def n_inputs(self):
        return sum(len(c[2]) for c in self.changed)

    @property
    def n_values(self):
        return sum(len(c[3]) for c in self.changed)

    @property
    def n_formulas(self):
        return sum(len(c[4]) for c in self.changed)

    def is_empty(self):
        cols = self.col_align
        col_changes = cols and (cols.inserted or cols.deleted or cols.moved)
        return not (self.inserted or self.deleted or self.moved or self.changed or col_changes)


@dataclass
class WorkbookDiff:
    old: LoadedBook
    new: LoadedBook
    status: str
    common: list
    sheets: dict                      # name -> (vb, va, SheetDiff)
    added_sheets: list = field(default_factory=list)
    removed_sheets: list = field(default_factory=list)
    renamed_sheets: list = field(default_factory=list)
    diagnostics: dict = field(default_factory=dict)
    computed_formulas: bool = False
    errors: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# compare()                                                                    #
# --------------------------------------------------------------------------- #

def _aborted(old_name, new_name, errors):
    diag = {"errors": errors, "warnings": [], "caps_hit": [],
            "cached_value_coverage": {}, "format_coverage": 0.0, "deshift_failures": {}}
    return WorkbookDiff(None, None, "aborted", [], {}, diagnostics=diag,
                        errors=errors)


def compare(old_path, new_path, *, formulas=False, want_graph=True) -> WorkbookDiff:
    """Diff two workbooks. NEVER raises: load/format failures return status='aborted'."""
    errors = []
    books = {}
    for tag, p in (("old", old_path), ("new", new_path)):
        try:
            books[tag] = load_book(p)
        except WorkbookLoadError as e:
            errors.append({"code": e.code, "path": e.path, "detail": e.detail})
    if errors:
        wd = _aborted(Path(old_path).name, Path(new_path).name, errors)
        wd.old = books.get("old")
        wd.new = books.get("new")
        return wd

    old, new = books["old"], books["new"]
    warnings, caps_hit, deshift_failures = [], [], {}

    sheets_a, sheets_b = set(old.order), set(new.order)
    added = [s for s in new.order if s not in sheets_a]
    removed = [s for s in old.order if s not in sheets_b]
    common = [s for s in new.order if s in sheets_a]
    renamed = _detect_renames(old, new, added, removed)
    for r in renamed:
        common.append(r["new"])                        # diff renamed as common pair
        added.remove(r["new"]); removed.remove(r["old"])

    status = "ok"
    sheet_results = {}
    for name in common:
        old_name = next((r["old"] for r in renamed if r["new"] == name), name)
        va = SheetView(old.content.get(old_name, {}), old.values.get(old_name, {}),
                       old.formats.get(old_name, {}))
        vb = SheetView(new.content.get(name, {}), new.values.get(name, {}),
                       new.formats.get(name, {}))

        if len(vb.populated_rows()) > MAX_ROWS or len(vb.populated_cols()) > MAX_COLS:
            caps_hit.append({"sheet": name, "kind": "rows" if len(vb.populated_rows()) > MAX_ROWS else "cols",
                             "limit": MAX_ROWS, "populated": len(vb.populated_rows()),
                             "action": "structural diff skipped for sheet"})
            status = "degraded"
            sheet_results[name] = (vb, va, SheetDiff(name, col_align=align_columns(va, vb)))
            continue

        ca = align_columns(va, vb)
        ins, dele, moved, changed, ca, blocks, cosmetic = diff_sheet(
            va, vb, want_formulas=formulas, col_align=ca)
        sd = SheetDiff(name, ins, dele, moved, blocks, changed, ca, cosmetic)
        sheet_results[name] = (vb, va, sd)

        cov = new.coverage.get(name, 1.0)
        if cov < 0.5:
            warnings.append({"code": "LOW_CACHE_COVERAGE", "sheet": name,
                             "detail": f"only {cov:.0%} of formula cells have cached values; "
                                       "open + save in Excel for reliable alignment.",
                             "severity": "high" if cov < 0.1 else "medium"})
        if ca.confidence == "low" and (ca.inserted or ca.deleted or len(vb.populated_cols()) != len(va.populated_cols())):
            warnings.append({"code": "LOW_COL_CONFIDENCE", "sheet": name,
                             "detail": "column structure uncertain; compared row-only.",
                             "severity": "medium"})
        df = sum(1 for (_ra, _rb, _i, _v, fs) in changed for f in fs if not f.get("deshift_ok", True))
        if df:
            deshift_failures[name] = df

    cosmetic_suppressed = {n: wd_sd.cosmetic_formula_shifts
                           for n, (_vb, _va, wd_sd) in sheet_results.items()
                           if wd_sd.cosmetic_formula_shifts}
    diagnostics = {
        "errors": [], "warnings": warnings, "caps_hit": caps_hit,
        "cached_value_coverage": {n: round(new.coverage.get(n, 1.0), 3) for n in common},
        "format_coverage": round(new.fmt_known / new.fmt_cells, 3) if new.fmt_cells else 0.0,
        "deshift_failures": deshift_failures,
        "formula_shifts_suppressed": cosmetic_suppressed,
    }
    wd = WorkbookDiff(old, new, status, common, sheet_results, added, removed,
                      renamed, diagnostics, computed_formulas=formulas)
    wd._want_graph = want_graph
    return wd


def _detect_renames(old, new, added, removed):
    """Pair an added+removed sheet whose row-label fingerprints are very similar."""
    def fp(book, name):
        vmap = book.values.get(name, {})
        labels = set()
        for (r, c), v in vmap.items():
            if c <= 3 and isinstance(v, str) and v.strip():
                labels.add(v.strip().lower())
        return labels
    renamed = []
    used_old = set()
    for nb in added:
        fnb = fp(new, nb)
        if len(fnb) < 5:
            continue
        best, best_sim = None, 0.0
        for ob in removed:
            if ob in used_old:
                continue
            fob = fp(old, ob)
            if len(fob) < 5:
                continue
            inter = len(fnb & fob)
            sim = inter / max(1, len(fnb | fob))
            if sim > best_sim:
                best, best_sim = ob, sim
        if best and best_sim >= 0.7:
            renamed.append({"old": best, "new": nb, "similarity": round(best_sim, 3)})
            used_old.add(best)
    return renamed


# --------------------------------------------------------------------------- #
# JSON emission (the ground-truth fact graph)                                  #
# --------------------------------------------------------------------------- #

def _unit_fields(view, r, c, number_format):
    kind, sym, scale = classify_unit(number_format)
    return {"number_format": number_format, "unit_kind": kind,
            "currency": sym, "scale": scale}


def _delta(old_raw, new_raw, old_kind, new_kind):
    out = {}
    if not (isinstance(old_raw, (int, float)) and isinstance(new_raw, (int, float))
            and not isinstance(old_raw, bool) and not isinstance(new_raw, bool)):
        return out
    if old_kind != new_kind and "unknown" not in (old_kind, new_kind):
        out["delta_display"] = "(unit changed)"        # mixing %, $ etc. is meaningless
        return out
    da = new_raw - old_raw
    out["delta_abs"] = da
    kind = new_kind
    if kind == "percent":
        out["delta_pp"] = da * 100
        out["delta_display"] = f"{da * 100:+.2f}pp"
    elif old_raw not in (0, 0.0):
        out["delta_pct"] = da / abs(old_raw)
        out["delta_display"] = f"{da / abs(old_raw) * 100:+.1f}%"
    else:
        out["delta_display"] = f"+{render_display(da)}" if da >= 0 else render_display(da)
    return out


def _role_weight(role):
    return {"output": 1.0, "total": 0.9, "input": 0.8, "driver": 0.8,
            "intermediate": 0.4, "header": 0.1, "spacer": 0.1}.get(role, 0.3)


def _materiality(old_raw, new_raw, role, line_scale):
    if not (isinstance(old_raw, (int, float)) and isinstance(new_raw, (int, float))):
        return round(0.2 * _role_weight(role), 4)
    mag = abs(new_raw - old_raw)
    norm = mag / line_scale if line_scale else mag
    return round(min(1.0, norm) * _role_weight(role), 4)


def emit_json(wd: WorkbookDiff) -> dict:
    """Serialise the diff into the deterministic, byte-stable fact graph."""
    if wd.status == "aborted":
        return {
            "schema_version": SCHEMA_VERSION, "engine_version": ENGINE_VERSION,
            "status": "aborted",
            "old": {"name": wd.old.name, "sha256": wd.old.sha256} if wd.old else None,
            "new": {"name": wd.new.name, "sha256": wd.new.sha256} if wd.new else None,
            "diagnostics": wd.diagnostics, "totals": {}, "sheets": [],
            "headline_metrics": [], "causal_links": [], "named_ranges": {},
        }

    old_sha, new_sha = wd.old.sha256, wd.new.sha256
    run_id = "wbdiff2-" + hashlib.sha256(
        (old_sha + new_sha + SCHEMA_VERSION + ENGINE_VERSION).encode()).hexdigest()[:16]

    raw_facts = []                                     # collected, then id-assigned

    def add(**kw):
        raw_facts.append(kw)

    # ---- per-sheet facts (structural + cell) ----
    sheet_blocks = {}
    for name in wd.common:
        vb, va, sd = wd.sheets[name]
        ca = sd.col_align
        # column structural facts
        for nc in ca.inserted:
            add(sheet=name, type="col_inserted", tier="structure", new_col=nc,
                label=vb.col_header(nc), sort=(name, "1col", nc))
        for oc in ca.deleted:
            add(sheet=name, type="col_deleted", tier="structure", old_col=oc,
                label=va.col_header(oc), sort=(name, "1col", oc))
        for oc, nc in ca.moved:
            add(sheet=name, type="col_moved", tier="structure", old_col=oc, new_col=nc,
                label=vb.col_header(nc), sort=(name, "1col", nc))
        # row structural facts (collapse block members)
        block_members = {m for blk in sd.moved_blocks for m in blk["members"]}
        for blk in sd.moved_blocks:
            lab = vb.label(blk["new_rows"][0]) or va.label(blk["old_rows"][0])
            add(sheet=name, type="block_moved", tier="structure",
                old_rows=blk["old_rows"], new_rows=blk["new_rows"], delta=blk["delta"],
                n_rows=blk["n_rows"], label=lab, sort=(name, "2row", blk["new_rows"][0]))
        for ra, rb in sd.moved:
            if (ra, rb) in block_members:
                continue
            add(sheet=name, type="row_moved", tier="structure", old_row=ra, new_row=rb,
                delta=rb - ra, label=vb.label(rb) or va.label(ra), sort=(name, "2row", rb))
        for rb in sd.inserted:
            role, rconf = vb.classify_role(rb)
            add(sheet=name, type="row_inserted", tier="structure", new_row=rb,
                label=vb.label(rb), role=role, preview=_row_preview(vb, rb),
                sort=(name, "2row", rb))
        for ra in sd.deleted:
            role, rconf = va.classify_role(ra)
            add(sheet=name, type="row_deleted", tier="structure", old_row=ra,
                label=va.label(ra), role=role, preview=_row_preview(va, ra),
                sort=(name, "2row", ra))
        # cell facts
        for (ra, rb, inputs, values, formulas) in sd.changed:
            role, rconf = vb.classify_role(rb)
            label = vb.label(rb) or va.label(ra)
            line_scale = _line_scale(vb, rb)
            for it in inputs:
                oc, nc = it["old_col"], it["new_col"]
                nf = vb.fmt_at(rb, nc) or va.fmt_at(ra, oc)
                u = _unit_fields(vb, rb, nc, nf)
                old_kind = classify_unit(va.fmt_at(ra, oc))[0]
                add(sheet=name, type="input", tier="cause", role=role, role_confidence=rconf,
                    old_addr=addr(ra, oc), new_addr=addr(rb, nc), label=label,
                    old_raw=it["old"], new_raw=it["new"],
                    old_display=render_display(it["old"], va.fmt_at(ra, oc)),
                    new_display=render_display(it["new"], nf),
                    **u, **_delta(it["old"], it["new"], old_kind, u["unit_kind"]),
                    _mat=_materiality(it["old"], it["new"], role, line_scale),
                    sort=(name, "3input", rb, nc))
            for it in formulas:
                oc, nc = it["old_col"], it["new_col"]
                add(sheet=name, type="formula", tier="cause", role=role, role_confidence=rconf,
                    old_addr=addr(ra, oc), new_addr=addr(rb, nc), label=label,
                    old_formula=it["old"], new_formula=it["new"], deshift_ok=it["deshift_ok"],
                    confidence="high" if it["deshift_ok"] else "low",
                    _mat=0.6 * _role_weight(role), sort=(name, "4formula", rb, nc))
            for it in values:
                oc, nc = it["old_col"], it["new_col"]
                nf = vb.fmt_at(rb, nc) or va.fmt_at(ra, oc)
                u = _unit_fields(vb, rb, nc, nf)
                old_kind = classify_unit(va.fmt_at(ra, oc))[0]
                add(sheet=name, type="value", tier="effect", role=role, role_confidence=rconf,
                    old_addr=addr(ra, oc), new_addr=addr(rb, nc), label=label,
                    old_raw=it["old"], new_raw=it["new"],
                    old_display=render_display(it["old"], va.fmt_at(ra, oc)),
                    new_display=render_display(it["new"], nf),
                    **u, **_delta(it["old"], it["new"], old_kind, u["unit_kind"]),
                    _mat=_materiality(it["old"], it["new"], role, line_scale),
                    sort=(name, "5value", rb, nc))

    # ---- cap value (effect) facts to the most material; the narrative cites
    #      headlines for effects, so uncapped ripple is not needed for grounding.
    VALUE_CAP = 3000
    value_facts = [f for f in raw_facts if f.get("type") == "value"]
    true_value_count = len(value_facts)
    values_omitted = 0
    if true_value_count > VALUE_CAP:
        # NEVER cap a headline-relevant effect: summary-sheet cells and
        # output/total rows are where headlines source their numbers.
        must = [f for f in value_facts
                if _SUMMARY_SHEET_RE.search(f["sheet"]) or f.get("role") in ("output", "total")]
        must_ids = set(id(f) for f in must)
        rest = sorted((f for f in value_facts if id(f) not in must_ids),
                      key=lambda f: (-(f.get("_mat") or 0), str(f.get("sort", ()))))
        keep = must_ids | set(id(f) for f in rest[:max(0, VALUE_CAP - len(must))])
        values_omitted = true_value_count - len(keep)
        raw_facts = [f for f in raw_facts if f.get("type") != "value" or id(f) in keep]

    # ---- deterministic id assignment ----
    sheet_order = {n: i for i, n in enumerate(wd.new.order)}
    raw_facts.sort(key=lambda f: (sheet_order.get(f["sheet"], 999),) + tuple(map(str, f.get("sort", ()))))
    facts = []
    addr_to_fact = {}                                  # (sheet,new_addr) -> id for value/effect
    for i, f in enumerate(raw_facts, 1):
        fid = f"F{i:04d}"
        f.pop("sort", None)
        mat = f.pop("_mat", None)
        f = {k: v for k, v in f.items() if v is not None}
        f["id"] = fid
        if mat is not None:
            f["materiality"] = mat
        facts.append(f)
        if f.get("new_addr"):
            addr_to_fact[(f["sheet"], f["new_addr"])] = fid

    # ---- named ranges ----
    named = _diff_named_ranges(wd.old, wd.new)

    # ---- headline metrics ----
    headlines = _extract_headlines(wd, facts)

    # ---- causal links (advisory) ----
    causal = []
    if getattr(wd, "_want_graph", True):
        causal = _attribute_causes(wd, facts, headlines, addr_to_fact)

    # ---- totals ----
    by_type = Counter(f["type"] for f in facts)
    totals = {
        "raw_naive_cell_changes": _naive_count(wd),
        "sheets_added": len(wd.added_sheets), "sheets_removed": len(wd.removed_sheets),
        "sheets_renamed": len(wd.renamed_sheets),
        "rows_inserted": by_type.get("row_inserted", 0), "rows_deleted": by_type.get("row_deleted", 0),
        "rows_moved": by_type.get("row_moved", 0), "blocks_moved": by_type.get("block_moved", 0),
        "cols_inserted": by_type.get("col_inserted", 0), "cols_deleted": by_type.get("col_deleted", 0),
        "cols_moved": by_type.get("col_moved", 0),
        "inputs_changed": by_type.get("input", 0), "formulas_changed": by_type.get("formula", 0),
        "values_rippled": true_value_count, "values_in_json": by_type.get("value", 0),
        "values_omitted_from_json": values_omitted,
        "named_ranges_changed": len(named["added"]) + len(named["removed"]) + len(named["changed"]),
    }

    changed_sheets = [n for n in wd.common if not wd.sheets[n][2].is_empty()]
    total_facts = len(facts)
    routing = {"total_facts": total_facts, "changed_sheets": len(changed_sheets),
               "recommend": "inline" if (total_facts < 150 and len(changed_sheets) <= 3) else "fanout_per_sheet"}

    # ---- nested sheet blocks ----
    facts_by_sheet = defaultdict(list)
    for f in facts:
        facts_by_sheet[f["sheet"]].append(f)
    sheets_json = []
    for name in wd.common:
        vb, va, sd = wd.sheets[name]
        ca = sd.col_align
        sheets_json.append({
            "name": name,
            "alignment": {"method": ca.method if ca else "none",
                          "col_confidence": ca.confidence if ca else "n/a",
                          "row_aligned": True},
            "columns": {
                "inserted": [{"new_index": c, "header": vb.col_header(c)} for c in (ca.inserted if ca else [])],
                "deleted": [{"old_index": c, "header": va.col_header(c)} for c in (ca.deleted if ca else [])],
                "moved": [{"old_index": o, "new_index": n, "header": vb.col_header(n)} for o, n in (ca.moved if ca else [])],
            },
            "fact_count": len(facts_by_sheet[name]),
            "facts": facts_by_sheet[name],
        })

    return {
        "schema_version": SCHEMA_VERSION,
        "diff_run_id": run_id,
        "engine_version": ENGINE_VERSION,
        "status": wd.status,
        "old": {"name": wd.old.name, "sha256": old_sha},
        "new": {"name": wd.new.name, "sha256": new_sha},
        "diagnostics": wd.diagnostics,
        "totals": totals,
        "routing": routing,
        "sheets_meta": {"old_order": wd.old.order, "new_order": wd.new.order,
                        "reordered": [s for s in wd.old.order if s in set(wd.new.order)]
                                     != [s for s in wd.new.order if s in set(wd.old.order)]},
        "renamed_sheets": wd.renamed_sheets,
        "sheets_added": wd.added_sheets,
        "sheets_removed": wd.removed_sheets,
        "named_ranges": named,
        "headline_metrics": headlines,
        "causal_links": causal,
        "sheets": sheets_json,
    }


def _line_scale(view, r):
    vals = [abs(v) for v in view.row_values(r).values()
            if isinstance(v, (int, float)) and not isinstance(v, bool) and isfinite(v)]
    return max(vals) if vals else 0


def _row_preview(view, r, max_cells=6):
    vals = view.row_values(r)
    parts = []
    for c in sorted(vals)[:max_cells]:
        parts.append(render_display(vals[c], view.fmt_at(r, c)))
    return parts


def _naive_count(wd):
    n = 0
    for name in wd.common:
        vb, va, _sd = wd.sheets[name]
        ca = set(va.content) | set(va.values)
        cb = set(vb.content) | set(vb.values)
        for k in ca | cb:
            if va.content.get(k) != vb.content.get(k) or _values_differ(
                    va.values.get(k), vb.values.get(k)):
                n += 1
    return n


def _diff_named_ranges(old, new):
    added, removed, changed = [], [], []
    oa, na = old.defined_names, new.defined_names
    for k in sorted(set(na) - set(oa)):
        added.append({"name": k, "refers_to": na[k]})
    for k in sorted(set(oa) - set(na)):
        removed.append({"name": k, "refers_to": oa[k]})
    for k in sorted(set(oa) & set(na)):
        if (oa[k] or "") != (na[k] or ""):
            changed.append({"name": k, "old_refers_to": oa[k], "new_refers_to": na[k]})
    return {"added": added, "removed": removed, "changed": changed}


_SUMMARY_SHEET_RE = re.compile(r"summary|dashboard|output|overview|p\s*&?\s*l|consol|results", re.I)


def _extract_headlines(wd, facts):
    """Closed, citable set of headline movements.

    Restricted to *summary* sheets (name-matched, else the first sheet): the
    intermediate calc sheets carry per-period 'output-lexicon' rows (CAC, ARPU)
    that are not the model's headline P&L. The as-of column is the rightmost of
    the first contiguous same-unit numeric run -- the annual/total column, not a
    monthly column or a scaled display dupe. Ranked by a unit-aware, degeneracy-
    damped materiality so a $20m revenue move outranks a +5000% ratio blip."""
    fact_by_addr = {(f["sheet"], f.get("new_addr")): f for f in facts if f.get("new_addr")}
    summary_sheets = [n for n in wd.common if _SUMMARY_SHEET_RE.search(n)]
    if not summary_sheets:
        summary_sheets = wd.common[:1]

    candidates = []
    for name in summary_sheets:
        vb, va, sd = wd.sheets[name]
        ca = sd.col_align
        common_new = sorted(set(ca.old_to_new.values())) if ca else sorted(vb.populated_cols())
        new_to_old = {n: o for o, n in ca.old_to_new.items()} if ca else {}
        rev_scale = _sheet_revenue_scale(vb, common_new)
        for (ra, rb, inputs, values, formulas) in sd.changed:
            role, rconf = vb.classify_role(rb)
            if role not in ("output", "total") or rconf == "low":
                continue
            label = vb.label(rb) or va.label(ra)
            if not label:
                continue
            nc = _as_of_col(vb, rb, common_new)
            if nc is None:
                continue
            oc = new_to_old.get(nc, nc)
            old_raw = va.row_values(ra).get(oc)
            new_raw = vb.row_values(rb).get(nc)
            if not (isinstance(old_raw, (int, float)) and isinstance(new_raw, (int, float))
                    and not isinstance(old_raw, bool) and not isinstance(new_raw, bool)):
                continue
            if not _values_differ(old_raw, new_raw):
                continue
            nf = vb.fmt_at(rb, nc)
            kind, sym, scale = classify_unit(nf)
            src = fact_by_addr.get((name, addr(rb, nc)))
            candidates.append({
                "sheet": name, "label": label, "role": role, "confidence": rconf,
                "as_of_col_header": vb.col_header(nc) or get_column_letter(nc),
                "unit_kind": kind, "old_raw": old_raw, "new_raw": new_raw,
                "old_display": render_display(old_raw, nf), "new_display": render_display(new_raw, nf),
                **_delta(old_raw, new_raw, kind, kind),
                "source_fact_id": src["id"] if src else None,
                "_mat": _headline_materiality(old_raw, new_raw, kind, role, rev_scale),
                "_key": (name, label.lower())})

    best = {}
    for c in candidates:
        k = c["_key"]
        if k not in best or c["_mat"] > best[k]["_mat"]:
            best[k] = c
    headlines = sorted(best.values(), key=lambda h: -h["_mat"])
    for h in headlines:
        h.pop("_mat", None); h.pop("_key", None)
    for i, h in enumerate(headlines[:20], 1):
        h["id"] = f"H{i:02d}"
    return [h for h in headlines if "id" in h]


def _sheet_revenue_scale(view, cols):
    """Largest as-of magnitude among output/total rows -- the denominator that
    keeps headline ranking comparable across lines (revenue dominates)."""
    scale = 0.0
    for r in view.populated_rows():
        role, conf = view.classify_role(r)
        if role not in ("output", "total"):
            continue
        nc = _as_of_col(view, r, cols)
        if nc is None:
            continue
        v = view.row_values(r).get(nc)
        if isinstance(v, (int, float)) and not isinstance(v, bool) and isfinite(v):
            scale = max(scale, abs(v))
    return scale


def _headline_materiality(old_raw, new_raw, kind, role, rev_scale):
    if kind == "percent":
        base = abs(new_raw - old_raw) * 100 / 5.0       # 5pp move == max
    elif rev_scale:
        base = abs(new_raw - old_raw) / rev_scale
    else:
        base = 0.0
    return round(min(1.0, base) * _role_weight(role), 5)


def _as_of_col(view, r, cols):
    """As-of column = rightmost of the FIRST contiguous run of same-unit numeric
    columns (the primary period block / annual total), preferring a column whose
    header explicitly marks a total/year. Avoids monthly tails and $m display
    dupes that sit to the right of the real total."""
    rv = view.row_values(r)
    numeric = [c for c in cols
               if isinstance(rv.get(c), (int, float)) and not isinstance(rv.get(c), bool)
               and isfinite(rv.get(c))]
    if not numeric:
        return None
    for c in reversed(numeric):                          # explicit total/year header wins
        h = view.col_header(c).lower()
        if re.search(r"\b(total|annual|fy\d*|full[\s-]*year|cumulative|20\d\d)\b", h):
            return c
    run = [numeric[0]]                                   # else first contiguous same-unit run
    for prev, c in zip(numeric, numeric[1:]):
        same_unit = classify_unit(view.fmt_at(r, c))[0] == classify_unit(view.fmt_at(r, prev))[0]
        if c == prev + 1 and same_unit:
            run.append(c)
        else:
            break
    return run[-1]


def _attribute_causes(wd, facts, headlines, addr_to_fact):
    """Advisory causal links: a cause cell whose dependents (in the NEW formula
    graph) include changed/headline cells. Confidence-tagged; never asserted."""
    deps = build_dependents(wd.new.content, wd.new.defined_names)
    cause_facts = [f for f in facts if f["tier"] == "cause" and f.get("new_addr")]
    if not cause_facts:
        return []
    # index effect/headline cells by (sheet, r, c)
    effect_nodes = {}
    for f in facts:
        if f["tier"] == "effect" and f.get("new_addr"):
            try:
                col, row = coordinate_from_string(f["new_addr"])
                effect_nodes[(f["sheet"], row, column_index_from_string(col))] = f["id"]
            except Exception:
                pass
    headline_nodes = {}
    for h in headlines:
        if h.get("source_fact_id"):
            sf = next((f for f in facts if f["id"] == h["source_fact_id"]), None)
            if sf and sf.get("new_addr"):
                try:
                    col, row = coordinate_from_string(sf["new_addr"])
                    headline_nodes[(sf["sheet"], row, column_index_from_string(col))] = h["id"]
                except Exception:
                    pass
    links = []
    for cf in cause_facts:
        try:
            col, row = coordinate_from_string(cf["new_addr"])
            start = (cf["sheet"], row, column_index_from_string(col))
        except Exception:
            continue
        reach = reachable(deps, start)
        hl_ids = sorted({headline_nodes[n] for n in reach if n in headline_nodes})
        if not hl_ids:
            continue                                    # only headline-reaching causes are narratable
        eff_ids = sorted({effect_nodes[n] for n in reach if n in effect_nodes})
        links.append({
            "cause_fact_id": cf["id"], "cause_label": cf.get("label"),
            "effect_fact_ids": eff_ids[:50], "effect_headline_ids": hl_ids,
            "method": "dep_graph_path", "path_proven": True, "confidence": "high"})
    for i, l in enumerate(sorted(links, key=lambda x: (-len(x["effect_headline_ids"]), x["cause_fact_id"])), 1):
        l["id"] = f"C{i:02d}"
    return links


# --------------------------------------------------------------------------- #
# Markdown rendering (capped human preview)                                    #
# --------------------------------------------------------------------------- #

def render_markdown(wd: WorkbookDiff, *, values=False, formulas=False, max_rows=0,
                    headlines=None, causal=None) -> str:
    L = []
    L.append("# Workbook diff (structural)")
    L.append("")
    if wd.status == "aborted":
        L.append("**Diff aborted — could not read the workbook(s):**")
        L.append("")
        for e in wd.errors:
            L.append(f"- `{e['code']}` {md_escape(e['path'])}: {md_escape(e['detail'])}")
        return "\n".join(L)

    L.append(f"- **OLD**: `{md_escape(wd.old.name)}`")
    L.append(f"- **NEW**: `{md_escape(wd.new.name)}`")
    if wd.status == "degraded":
        L.append("- _Status: DEGRADED — some sheets exceeded size caps (see diagnostics)._")
    if not values:
        L.append("- _Computed-value ripple omitted; pass `--values` to include it._")
    L.append("")

    diag = wd.diagnostics
    if diag.get("warnings") or diag.get("caps_hit"):
        L.append("## Diagnostics")
        L.append("")
        for w in diag.get("warnings", []):
            L.append(f"- ⚠️ `{w['code']}` [{w['sheet']}]: {md_escape(w['detail'])}")
        for c in diag.get("caps_hit", []):
            L.append(f"- ⚠️ size cap on `{c['sheet']}`: {c['action']} ({c['populated']} {c['kind']})")
        L.append("")

    if wd.added_sheets or wd.removed_sheets or wd.renamed_sheets:
        L.append("## Sheets")
        L.append("")
        for s in wd.added_sheets:
            L.append(f"- **added**: {md_escape(s)}")
        for s in wd.removed_sheets:
            L.append(f"- **removed**: {md_escape(s)}")
        for r in wd.renamed_sheets:
            L.append(f"- **renamed**: {md_escape(r['old'])} -> {md_escape(r['new'])} (similarity {r['similarity']})")
        L.append("")

    L.append("## Summary")
    L.append("")
    L.append("| Sheet | Cols +/- | Inserted | Deleted | Moved | Inputs | Values |"
             + (" Formulas |" if formulas else ""))
    L.append("|---|---|---|---|---|---|---|" + ("---|" if formulas else ""))
    for name in wd.common:
        vb, va, sd = wd.sheets[name]
        if sd.is_empty():
            continue
        ca = sd.col_align
        coln = f"+{len(ca.inserted)}/-{len(ca.deleted)}" if ca else "-"
        row = (f"| {md_escape(name)} | {coln} | {len(sd.inserted)} | {len(sd.deleted)} "
               f"| {len(sd.moved)} | {sd.n_inputs} | {sd.n_values} |")
        if formulas:
            row += f" {sd.n_formulas} |"
        L.append(row)
    L.append("")

    if headlines is None:
        headlines = _extract_headlines(wd, [])
    if headlines:
        L.append("## Headline impact")
        L.append("")
        L.append("| Metric | As of | Old | New | Change |")
        L.append("|---|---|---|---|---|")
        for h in headlines[:15]:
            L.append(f"| {md_escape(h['label'])} | {md_escape(h['as_of_col_header'])} | "
                     f"{md_escape(h['old_display'])} | {md_escape(h['new_display'])} | "
                     f"{md_escape(h.get('delta_display', ''))} |")
        L.append("")
        if causal:
            L.append("_Drivers (dependency-path proven):_ "
                     + "; ".join(f"{md_escape(c.get('cause_label') or c['cause_fact_id'])} -> "
                                 f"{', '.join(c['effect_headline_ids'])}" for c in causal[:6]))
            L.append("")

    L.append("## Detail")
    L.append("")
    any_detail = False
    for name in wd.common:
        vb, va, sd = wd.sheets[name]
        if sd.is_empty():
            continue
        any_detail = True
        L.append(f"### {md_escape(name)}")
        L.append("")
        ca = sd.col_align
        if ca and (ca.inserted or ca.deleted or ca.moved):
            L.append("**Columns**")
            L.append("")
            for c in ca.inserted:
                L.append(f"- inserted: **{md_escape(vb.col_header(c) or get_column_letter(c))}**")
            for c in ca.deleted:
                L.append(f"- deleted: **{md_escape(va.col_header(c) or get_column_letter(c))}**")
            for o, n in ca.moved:
                L.append(f"- moved: **{md_escape(vb.col_header(n) or get_column_letter(n))}**")
            L.append("")

        if sd.moved_blocks:
            L.append(f"**Blocks moved ({len(sd.moved_blocks)})**")
            L.append("")
            for blk in _cap(sd.moved_blocks, max_rows):
                lab = vb.label(blk["new_rows"][0]) or va.label(blk["old_rows"][0]) or "(block)"
                L.append(f"- **{md_escape(lab)}**: rows {blk['old_rows'][0]}-{blk['old_rows'][1]} "
                         f"-> {blk['new_rows'][0]}-{blk['new_rows'][1]} ({'down' if blk['delta']>0 else 'up'} {abs(blk['delta'])})")
            L.append("")

        block_members = {m for blk in sd.moved_blocks for m in blk["members"]}
        single_moves = [(ra, rb) for ra, rb in sd.moved if (ra, rb) not in block_members]
        if single_moves:
            L.append(f"**Rows moved ({len(single_moves)})**")
            L.append("")
            for ra, rb in _cap(single_moves, max_rows):
                lab = vb.label(rb) or va.label(ra) or "(no label)"
                d = rb - ra
                L.append(f"- **{md_escape(lab)}**: old row {ra} -> new row {rb} ({'down' if d>0 else 'up'} {abs(d)})")
            _more(L, single_moves, max_rows, "moved rows")
            L.append("")

        if sd.inserted:
            L.append(f"**Rows inserted ({len(sd.inserted)})**")
            L.append("")
            for rb in _cap(sd.inserted, max_rows):
                lab = vb.label(rb) or "(no label)"
                prev = ", ".join(md_escape(x) for x in _row_preview(vb, rb))
                L.append(f"- new row {rb}: **{md_escape(lab)}** [{prev}]")
            _more(L, sd.inserted, max_rows, "inserted rows")
            L.append("")

        if sd.deleted:
            L.append(f"**Rows deleted ({len(sd.deleted)})**")
            L.append("")
            for ra in _cap(sd.deleted, max_rows):
                lab = va.label(ra) or "(no label)"
                prev = ", ".join(md_escape(x) for x in _row_preview(va, ra))
                L.append(f"- old row {ra}: **{md_escape(lab)}** [{prev}]")
            _more(L, sd.deleted, max_rows, "deleted rows")
            L.append("")

        # inputs, ranked by materiality
        input_items = []
        for (ra, rb, inputs, _v, _f) in sd.changed:
            for it in inputs:
                input_items.append((ra, rb, it))
        if input_items:
            L.append("**Inputs changed**")
            L.append("")
            for ra, rb, it in _cap(input_items, max_rows):
                lab = vb.label(rb) or va.label(ra) or "(no label)"
                a = addr(rb, it["new_col"])
                od = render_display(it["old"], va.fmt_at(ra, it["old_col"]))
                nd = render_display(it["new"], vb.fmt_at(rb, it["new_col"]))
                L.append(f"- {a} **{md_escape(lab)}**: `{md_escape(od)}` => `{md_escape(nd)}`")
            _more(L, input_items, max_rows, "input changes")
            L.append("")

        if formulas:
            fitems = [(ra, rb, it) for (ra, rb, _i, _v, fs) in sd.changed for it in fs]
            if fitems:
                L.append("**Formulas changed** (shift-corrected)")
                L.append("")
                for ra, rb, it in _cap(fitems, max_rows):
                    lab = vb.label(rb) or va.label(ra) or "(no label)"
                    a = addr(rb, it["new_col"])
                    tag = "" if it["deshift_ok"] else " _(advisory: de-shift failed)_"
                    L.append(f"- {a} **{md_escape(lab)}**: `{md_escape(formula_text(it['old']))}` "
                             f"=> `{md_escape(formula_text(it['new']))}`{tag}")
                _more(L, fitems, max_rows, "formula changes")
                L.append("")

        if values:
            vitems = []
            for (ra, rb, _i, vs, _f) in sd.changed:
                if vs:
                    vitems.append((ra, rb, vs))
            if vitems:
                L.append("**Values moved** (computed results)")
                L.append("")
                for ra, rb, vs in _cap(vitems, max_rows):
                    lab = vb.label(rb) or va.label(ra) or "(no label)"
                    cells = ", ".join(
                        f"{addr(rb, it['new_col'])} {md_escape(render_display(it['old'], vb.fmt_at(rb, it['new_col'])))}"
                        f"=>{md_escape(render_display(it['new'], vb.fmt_at(rb, it['new_col'])))}"
                        for it in vs[:8])
                    extra = f" (+{len(vs)-8} more)" if len(vs) > 8 else ""
                    L.append(f"- row {rb} **{md_escape(lab)}**: {cells}{extra}")
                _more(L, vitems, max_rows, "value changes")
                L.append("")

    if not any_detail and not wd.added_sheets and not wd.removed_sheets:
        L.append("_No structural or content differences found._")
    return "\n".join(L)


def _cap(seq, n):
    return seq[:n] if n else seq


def _more(L, seq, n, noun):
    if n and len(seq) > n:
        L.append(f"- ... and {len(seq) - n} more {noun}")


# --------------------------------------------------------------------------- #
# Convenience                                                                  #
# --------------------------------------------------------------------------- #

def emit_summary_json(wd: WorkbookDiff) -> dict:
    """A small, agent-readable view of the fact graph: the full header,
    diagnostics, totals, routing, named ranges, headline_metrics and
    causal_links, plus only the CAUSE + STRUCTURAL facts per sheet (the value/
    effect ripple is dropped -- the narrative cites headlines for effects). This
    is what the skill reads to author; the full diff.json feeds the verifier."""
    j = emit_json(wd)
    if j.get("status") == "aborted":
        return j
    for s in j.get("sheets", []):
        kept = [f for f in s["facts"] if f.get("type") != "value"]
        s["value_facts_omitted"] = len(s["facts"]) - len(kept)
        s["facts"] = kept
    j["view"] = "summary"
    return j


def diff_to_markdown(old, new, *, formulas=False, values=False, max_rows=0):
    return render_markdown(compare(old, new, formulas=formulas),
                           values=values, formulas=formulas, max_rows=max_rows)


def diff_to_json(old, new, *, formulas=True):
    return emit_json(compare(old, new, formulas=formulas))
